from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8")


def clean_cell(value) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def main() -> int:
    if len(sys.argv) < 2:
        print("Usage: python extract_letters_xlsx.py <input.xlsx> [output.json]")
        return 1

    input_path = Path(sys.argv[1]).resolve()
    if not input_path.exists():
        print(f"Input file not found: {input_path}")
        return 1

    if len(sys.argv) >= 3:
        output_path = Path(sys.argv[2]).resolve()
    else:
        output_path = (
            Path(__file__).resolve().parent / "letters_from_xlsx_1404.json"
        )

    wb = load_workbook(input_path, data_only=True)
    ws = wb.active

    rows = []
    for row_idx in range(2, ws.max_row + 1):
        vals = [clean_cell(ws.cell(row=row_idx, column=col).value) for col in range(1, 16)]
        if not any(vals):
            continue

        rows.append(
            {
                "letter_no": vals[0],
                "letter_date": vals[1],
                "secretariat_no": vals[2],
                "secretariat_date": vals[3],
                "project_name": vals[4],
                "project_code": vals[5],
                "subject": vals[6],
                "kind_raw": vals[7],
                "from_name": vals[8],
                "to_name": vals[9],
                "org_name": vals[10],
                "related_doc": vals[11],
                "attachment_flag": vals[12],
                "attachment_desc": vals[13],
                "tag": vals[14],
            }
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    output_path.write_text(
        json.dumps({"rows": rows}, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    print(f"Input: {input_path}")
    print(f"Output: {output_path}")
    print(f"Rows: {len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
